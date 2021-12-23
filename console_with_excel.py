import openpyxl
import serial
import time
import os

def vlan_send_command(vlanID,vlanIP,vlanIPNetMask):
    global ser
    vlan_ID = str.encode('interface vlan ' + str(vlanID) + '\r\n')
    vlan_IP = str.encode('ip address ' + vlanIP + ' ')
    vlan_NetMask = str.encode(vlanIPNetMask + '\r\n')
    ser.write(vlan_ID)
    ser.write(vlan_IP + vlan_NetMask)



ser = serial.Serial(
    port = 'COM4',
    baudrate=9600,
    timeout=8
    )
input_data = ser.read(225)
input_data = input_data.decode("utf-8", "ignore")
ser.write(b'enable\r\n')
ser.write(b'config t\r\n')

if os.path.isfile('console.xlsx'):
    wb = openpyxl.load_workbook('console.xlsx')
else:
    print('please check the file "console.xlsx" was exits')

vlan_sheet = wb.worksheets[0]
vlan_row_count = vlan_sheet.max_row

for i in range(2,vlan_row_count+1):
    vlanID=vlan_sheet.cell(row=i, column=1).value
    vlanIP=vlan_sheet.cell(row=i, column=2).value
    vlanIPnetmask = vlan_sheet.cell(row=i, column=3).value
    vlan_ID = str.encode('interface vlan ' + str(vlanID) + '\r\n')
    vlan_IP = str.encode('ip address ' + vlanIP + ' ')
    vlan_NetMask = str.encode(vlanIPnetmask + '\r\n')
    ser.write(vlan_ID)
    ser.write(vlan_IP + vlan_NetMask)

int_sheet = wb.worksheets[1]
int_row_count = int_sheet.max_row

for i in range(2,int_max_row + 1):
    interface = int_sheet.cell(row=i,column=1).value
    int_mode = int_sheet.cell(row=i,column=2).value
    interface_PortID = str.encode('interface ' + interface + '\r\n')
    ser.write(interface_PortID)
    if 'no switch' in int_mode:
        ip = int_sheet.cell(row=i,column=3).value
        netmask = int_sheet.cell(row=i,column=4).value
        inter_IP = str.encode('ip address ' + ip + ' ' + netmask + '\r\n')
        ser.write(b'no switchport\r\n')
        ser.write(inter_IP)
    else:
        inter_mode = str.encode('switchport mode ' + int_mode +'\r\n')
        ser.write(inter_mode)

ser.write(b'end\r\n')
ser.write(b'wr\r\n')
ser.close()
